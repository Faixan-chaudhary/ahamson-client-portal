"""Quotation Excel export, stats, and branded professional PDF quotes."""
from __future__ import annotations

from datetime import UTC, datetime
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    HRFlowable,
    Image,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from sqlalchemy.orm import Session

from app.models import Quotation, User
from app.quotations import list_quotations

BACKEND_ROOT = Path(__file__).resolve().parent.parent
PROJECT_ROOT = BACKEND_ROOT.parent
LOGO_CANDIDATES = [
    BACKEND_ROOT / "assets" / "ahamson-logo-transparent.png",
    PROJECT_ROOT / "public" / "logos" / "ahamson-logo.png",
    PROJECT_ROOT / "dist" / "logos" / "ahamson-logo.png",
]

NAVY_DEEP = colors.HexColor("#06142A")
GOLD = colors.HexColor("#F7931E")
SLATE = colors.HexColor("#64748B")
LINE = colors.HexColor("#E2E8F0")
WHITE = colors.white

HEADERS = [
    "Quote #", "Phase", "Status", "Quote Date", "Sales Person", "Partner", "End User",
    "Country", "Brand", "Products", "Deal Value", "GP", "Contact", "Closure",
    "Probability", "Submission Date", "Closure Reason", "OEM", "Updated",
]


def export_quotations_excel(
    db: Session,
    search: str | None = None,
    status_filter: str | None = None,
    phase: str | None = None,
    user: User | None = None,
) -> bytes:
    result = list_quotations(db, search=search, status_filter=status_filter, phase=phase, user=user)
    wb = Workbook()
    ws = wb.active
    ws.title = "Quotations"
    header_fill = PatternFill("solid", fgColor="0B1F3A")
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    thin = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(1, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin

    data_font = Font(name="Calibri", size=10)
    for i, q in enumerate(result.items, start=2):
        values = [
            q.quote_number, q.phase, q.status, q.quotation_date, q.sales_person,
            q.partner, q.end_user, q.country, q.brand, q.products,
            q.deal_value, q.gp_value, q.contact_person, q.closure_date,
            q.probability,
            q.quotation_submission_date or q.formal_submission_date,
            q.closure_reason, q.oem, q.updated_at,
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(i, col, val or "")
            cell.font = data_font
            cell.border = thin
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for col in range(1, len(HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.column_dimensions["J"].width = 28
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def quotation_stats(db: Session, user=None) -> dict:
    from app.quotation_status import (
        FINANCE_QUEUE_STATUSES,
        SALES_HEAD_QUEUE_STATUSES,
        SALES_QUEUE_STATUSES,
        allowed_actions,
    )

    rows = db.query(Quotation).all()
    by_status: dict[str, int] = {}
    by_phase = {"budgetary": 0, "formal": 0}
    submitted = 0
    lost = 0
    won_closed = 0
    open_active = 0
    pending_finance = 0
    pending_sales_head = 0
    pending_sales = 0
    my_queue = 0
    child_parent_ids: set[int] | None = None

    for row in rows:
        st = row.status or "Unknown"
        by_status[st] = by_status.get(st, 0) + 1
        if row.phase in by_phase:
            by_phase[row.phase] += 1
        low = st.lower()
        if "submitted" in low or "placed" in low:
            submitted += 1
        if "lost" in low or "not approved" in low:
            lost += 1
        elif "closed" in low:
            won_closed += 1
        else:
            open_active += 1
        if st in FINANCE_QUEUE_STATUSES:
            pending_finance += 1
        if st in SALES_HEAD_QUEUE_STATUSES:
            pending_sales_head += 1
        if st in SALES_QUEUE_STATUSES:
            pending_sales += 1
        if user is not None:
            if child_parent_ids is None:
                child_parent_ids = {
                    r.parent_quote_id
                    for r in rows
                    if r.parent_quote_id is not None
                }
            has_child = bool(row.id in child_parent_ids)
            if allowed_actions(row, user, has_formal_child=has_child):
                my_queue += 1

    stages = [
        {"status": k, "count": v}
        for k, v in sorted(by_status.items(), key=lambda kv: (-kv[1], kv[0]))
    ]
    return {
        "total": len(rows),
        "open": open_active,
        "submitted": submitted,
        "lost": lost,
        "closed": won_closed,
        "pending_finance": pending_finance,
        "pending_sales_head": pending_sales_head,
        "pending_sales": pending_sales,
        "my_queue": my_queue,
        "by_phase": by_phase,
        "by_status": stages,
    }


def _logo_path() -> Path | None:
    for path in LOGO_CANDIDATES:
        if path.is_file():
            return path
    return None


def _txt(value: object | None, fallback: str = "—") -> str:
    text = str(value or "").strip()
    return text if text else fallback


def _phase_title(phase: str | None) -> str:
    p = (phase or "").strip().lower()
    if p == "formal":
        return "Formal Quotation"
    return "Budgetary Quotation"


def _is_formal(phase: str | None) -> bool:
    return (phase or "").strip().lower() == "formal"


def pdf_filename_for(row: Quotation) -> str:
    """Stable, phase-clear download name."""
    number = (row.quote_number or f"quote-{row.id}").replace("/", "-").strip()
    kind = "Formal" if _is_formal(row.phase) else "Budgetary"
    return f"AHamson-{kind}-{number}.pdf"


def _format_date(value: str | None, *, fallback: str | None = None) -> str:
    text = (value or "").strip()
    candidates = [text]
    if fallback:
        candidates.append(fallback.strip())
    now_year = datetime.now(UTC).year
    for raw in candidates:
        if not raw:
            continue
        try:
            if "T" in raw:
                dt = datetime.fromisoformat(raw.replace("Z", "+00:00"))
            else:
                dt = datetime.fromisoformat(f"{raw[:10]}T00:00:00")
            # Ignore junk/autofill years (e.g. 1970 / 1980) on PDF output.
            if 2000 <= dt.year <= now_year + 1:
                return dt.strftime("%d %b %Y")
        except ValueError:
            continue
    return "—"


def build_quotation_pdf(row: Quotation, *, linked_quote_number: str | None = None) -> bytes:
    """Corporate AHamson quotation PDF — phase-specific Budgetary / Formal document."""
    buf = BytesIO()
    # Extra bottom margin reserved for fixed canvas footer (disclaimer always at page bottom).
    footer_h = 22 * mm
    is_formal = _is_formal(row.phase)
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        topMargin=12 * mm,
        bottomMargin=footer_h + 6 * mm,
        title=f"{row.quote_number} — AHamson {_phase_title(row.phase)}",
        author="AHamson",
    )
    width = A4[0] - doc.leftMargin - doc.rightMargin
    generated_at = datetime.now(UTC).strftime("%d %b %Y · %H:%M UTC")
    phase_title = _phase_title(row.phase)
    doc_intro = (
        "Formal quotation issued by AHamson for System Integrator / Partner review and order progression."
        if is_formal
        else "Budgetary quotation issued by AHamson for preliminary System Integrator / Partner review. "
             "Figures are indicative until a Formal Quotation is issued."
    )

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        "DocTitle", parent=styles["Normal"], fontName="Helvetica-Bold",
        fontSize=16, textColor=NAVY_DEEP, leading=20, spaceAfter=0, alignment=TA_LEFT,
    ))
    styles.add(ParagraphStyle(
        "DocSub", parent=styles["Normal"], fontName="Helvetica",
        fontSize=8.5, textColor=SLATE, leading=11, spaceAfter=0,
    ))
    styles.add(ParagraphStyle(
        "Section", parent=styles["Normal"], fontName="Helvetica-Bold",
        fontSize=8, textColor=NAVY_DEEP, leading=10,
        spaceBefore=7 * mm, spaceAfter=2.5 * mm,
        letterSpacing=0.6,
    ))
    styles.add(ParagraphStyle(
        "Label", parent=styles["Normal"], fontName="Helvetica",
        fontSize=7, textColor=SLATE, leading=9, spaceAfter=1,
    ))
    styles.add(ParagraphStyle(
        "Value", parent=styles["Normal"], fontName="Helvetica-Bold",
        fontSize=9, textColor=NAVY_DEEP, leading=11,
    ))
    styles.add(ParagraphStyle(
        "Body", parent=styles["Normal"], fontName="Helvetica",
        fontSize=9, textColor=NAVY_DEEP, leading=12,
    ))
    styles.add(ParagraphStyle(
        "MetaRight", parent=styles["Normal"], fontName="Helvetica",
        fontSize=8, textColor=colors.HexColor("#CBD5E1"), leading=10, alignment=TA_RIGHT,
    ))
    styles.add(ParagraphStyle(
        "MetaRightBold", parent=styles["Normal"], fontName="Helvetica-Bold",
        fontSize=10, textColor=WHITE, leading=12, alignment=TA_RIGHT,
    ))
    styles.add(ParagraphStyle(
        "StatusText", parent=styles["Normal"], fontName="Helvetica",
        fontSize=8.5, textColor=NAVY_DEEP, leading=11, alignment=TA_RIGHT,
    ))

    story: list = []

    logo = _logo_path()
    logo_cell: object = Paragraph(
        "<font color='white'><b>AHAMSON</b></font>",
        ParagraphStyle("LogoFallback", fontName="Helvetica-Bold", fontSize=14, textColor=WHITE, leading=16),
    )
    if logo:
        logo_cell = Image(str(logo), width=44 * mm, height=7.2 * mm)

    header_right = Table(
        [
            [Paragraph(phase_title.upper(), styles["MetaRightBold"])],
            [Paragraph(_txt(row.quote_number), styles["MetaRight"])],
            [Paragraph(
                _format_date(
                    row.quotation_date,
                    fallback=row.created_at.date().isoformat() if row.created_at else None,
                ),
                styles["MetaRight"],
            )],
        ],
        colWidths=[width * 0.42],
    )
    header_right.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), NAVY_DEEP),
        ("ALIGN", (0, 0), (-1, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 1),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
    ]))
    header = Table([[logo_cell, header_right]], colWidths=[width * 0.55, width * 0.45])
    header.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), NAVY_DEEP),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (0, 0), 14),
        ("RIGHTPADDING", (1, 0), (1, 0), 14),
        ("TOPPADDING", (0, 0), (-1, -1), 12),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
    ]))
    story.append(header)
    # Thin brand accent — one line only, not decorative chrome
    story.append(HRFlowable(width="100%", thickness=2, color=GOLD, spaceBefore=0, spaceAfter=6 * mm))

    title_row = Table(
        [[
            Paragraph(phase_title, styles["DocTitle"]),
            Paragraph(_txt(row.status), styles["StatusText"]),
        ]],
        colWidths=[width * 0.58, width * 0.42],
    )
    title_row.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "BOTTOM"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    story.append(title_row)
    story.append(Paragraph(doc_intro, styles["DocSub"]))
    if linked_quote_number:
        story.append(Paragraph(
            (
                f"Linked Budgetary Quotation: <b>{linked_quote_number}</b>"
                if is_formal
                else f"Linked Formal Quotation: <b>{linked_quote_number}</b>"
            ),
            styles["DocSub"],
        ))
    story.append(Spacer(1, 3 * mm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#94A3B8"), spaceAfter=1 * mm))

    def field(label: str, value: object | None) -> list:
        return [
            Paragraph(label.upper(), styles["Label"]),
            Paragraph(_txt(value), styles["Value"]),
        ]

    def section_table(rows: list, *, emphasize_first_row: bool = False) -> Table:
        table = Table(rows, colWidths=[width / 2, width / 2])
        style_cmds = [
            ("BACKGROUND", (0, 0), (-1, -1), WHITE),
            ("BOX", (0, 0), (-1, -1), 0.6, NAVY_DEEP),
            ("INNERGRID", (0, 0), (-1, -1), 0.35, LINE),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]
        if emphasize_first_row:
            style_cmds.append(("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F1F5F9")))
        table.setStyle(TableStyle(style_cmds))
        return table

    story.append(Paragraph("1. DOCUMENT CONTROL", styles["Section"]))
    story.append(section_table([
        [
            field("Document Type", phase_title),
            field("Quote Number", row.quote_number),
        ],
        [
            field(
                "Quote Date",
                _format_date(
                    row.quotation_date,
                    fallback=row.created_at.date().isoformat() if row.created_at else None,
                ),
            ),
            field("Workflow Status", row.status),
        ],
        [
            field("Linked Quote", linked_quote_number or "—"),
            field(
                "Submission Date",
                row.formal_submission_date if is_formal else row.quotation_submission_date,
            ),
        ],
    ]))

    story.append(Paragraph("2. PARTIES", styles["Section"]))
    story.append(section_table([
        [field("Partner (SI)", row.partner), field("End User", row.end_user)],
        [field("Sales Person", row.sales_person), field("Contact Person", row.contact_person)],
        [field("Country", row.country), field("Brand", row.brand)],
    ]))

    story.append(Paragraph(
        "3. COMMERCIAL SUMMARY — FORMAL" if is_formal else "3. COMMERCIAL SUMMARY — BUDGETARY",
        styles["Section"],
    ))
    story.append(section_table(
        [
            [field("Deal Value", row.deal_value), field("Gross Profit (GP)", row.gp_value)],
            [field("Probability", row.probability), field("Expected Closure", row.closure_date)],
            [
                field(
                    "Formal Submission" if is_formal else "Budgetary Submission",
                    row.formal_submission_date if is_formal else row.quotation_submission_date,
                ),
                field(
                    "Quote Date",
                    _format_date(
                        row.quotation_date,
                        fallback=row.created_at.date().isoformat() if row.created_at else None,
                    ),
                ),
            ],
        ],
        emphasize_first_row=True,
    ))

    story.append(Paragraph("4. PRODUCTS / SCOPE", styles["Section"]))
    products_box = Table(
        [[Paragraph(_txt(row.products).replace("\n", "<br/>"), styles["Body"])]],
        colWidths=[width],
    )
    products_box.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 0.6, NAVY_DEEP),
        ("LEFTPADDING", (0, 0), (-1, -1), 9),
        ("RIGHTPADDING", (0, 0), (-1, -1), 9),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))
    story.append(products_box)

    story.append(Paragraph("5. DETAILS / NOTES", styles["Section"]))
    details_box = Table(
        [[Paragraph(_txt(row.details).replace("\n", "<br/>"), styles["Body"])]],
        colWidths=[width],
    )
    details_box.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 0.6, LINE),
        ("LEFTPADDING", (0, 0), (-1, -1), 9),
        ("RIGHTPADDING", (0, 0), (-1, -1), 9),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))
    story.append(details_box)

    # Formal-only commercial / order block — keep Budgetary PDF clean.
    if is_formal and (row.oem or row.order_details or row.po_file_name or row.order_date):
        story.append(Paragraph("6. ORDER / OEM REFERENCE", styles["Section"]))
        story.append(section_table([
            [field("OEM", row.oem), field("PO Reference", row.po_file_name)],
            [field("Order Date", row.order_date), field("Expected Delivery", row.expected_delivery_date)],
            [field("Order Submission", row.order_submission_date), field("Order Remarks", row.order_remarks)],
        ]))
        if row.order_details:
            story.append(Spacer(1, 2 * mm))
            story.append(Paragraph(_txt(row.order_details).replace("\n", "<br/>"), styles["Body"]))

    if row.closure_reason:
        story.append(Paragraph("CLOSURE", styles["Section"]))
        story.append(Paragraph(
            f"<b>Reason:</b> {_txt(row.closure_reason)}"
            + (f"<br/><b>Remarks:</b> {_txt(row.closure_remarks)}" if row.closure_remarks else ""),
            styles["Body"],
        ))

    # Signature block stays with content (not the page footer).
    story.append(Spacer(1, 10 * mm))
    story.append(HRFlowable(width="100%", thickness=0.4, color=LINE, spaceAfter=4 * mm))
    prepared_label = "AHamson Sales — Formal Quotation" if is_formal else "AHamson Sales — Budgetary Quotation"
    sign = Table(
        [[
            Paragraph(
                f"Prepared by<br/><b>{prepared_label}</b><br/><font size='7' color='#64748B'>________________________</font>",
                styles["Body"],
            ),
            Paragraph(
                "For<br/><b>System Integrator / Partner</b><br/><font size='7' color='#64748B'>________________________</font>",
                styles["Body"],
            ),
        ]],
        colWidths=[width / 2, width / 2],
    )
    sign.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(sign)

    def _draw_page(canvas, _doc):
        """Disclaimer + meta always pinned to the physical bottom of every page."""
        canvas.saveState()
        left = 18 * mm
        right = A4[0] - 18 * mm
        y_top = footer_h

        canvas.setStrokeColor(NAVY_DEEP)
        canvas.setLineWidth(0.8)
        canvas.line(left, y_top, right, y_top)
        canvas.setStrokeColor(GOLD)
        canvas.setLineWidth(1.4)
        canvas.line(left, y_top - 1.6, right, y_top - 1.6)

        canvas.setFillColor(SLATE)
        canvas.setFont("Helvetica", 6.5)
        disclaimer = (
            (
                "This Formal Quotation is system-generated by the AHamson Client Portal. "
                "Commercial terms are subject to confirmation and internal approvals. "
                "For queries contact your AHamson sales representative."
            )
            if is_formal
            else (
                "This Budgetary Quotation is system-generated by the AHamson Client Portal. "
                "Figures are indicative and do not constitute a firm offer until a Formal Quotation is issued. "
                "For queries contact your AHamson sales representative."
            )
        )
        # Wrap disclaimer manually for canvas
        max_chars = 108
        words = disclaimer.split()
        lines: list[str] = []
        current = ""
        for word in words:
            trial = f"{current} {word}".strip()
            if len(trial) <= max_chars:
                current = trial
            else:
                lines.append(current)
                current = word
        if current:
            lines.append(current)

        text_y = y_top - 6 * mm
        for line in lines[:3]:
            canvas.drawCentredString(A4[0] / 2, text_y, line)
            text_y -= 3.2 * mm

        canvas.setFillColor(NAVY_DEEP)
        canvas.setFont("Helvetica", 7)
        canvas.drawString(left, 4.5 * mm, "AHAMSON")
        canvas.setFont("Helvetica", 6.5)
        canvas.setFillColor(SLATE)
        canvas.drawCentredString(A4[0] / 2, 4.5 * mm, f"Generated {generated_at}  ·  Confidential")
        canvas.drawRightString(right, 4.5 * mm, f"Page {_doc.page}")
        canvas.restoreState()

    doc.build(story, onFirstPage=_draw_page, onLaterPages=_draw_page)
    return buf.getvalue()
