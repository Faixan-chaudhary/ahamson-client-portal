"""AHamson sales pipeline — list/create/update/delete + Excel template export."""
from __future__ import annotations

import re
from copy import copy
from datetime import UTC, date, datetime
from io import BytesIO
from pathlib import Path

from fastapi import HTTPException, status
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.models import PipelineEntry, Quotation, User
from app.schemas import (
    PipelineEntryCreate,
    PipelineEntryOut,
    PipelineEntryUpdate,
    PipelineListResponse,
    PipelineSyncResponse,
)
from app.security import iso

BACKEND_ROOT = Path(__file__).resolve().parent.parent
PROJECT_ROOT = BACKEND_ROOT.parent
PIPELINE_TEMPLATE = PROJECT_ROOT / "public" / "forms" / "A Hamson - Pipeline Sheet.xlsx"

HEADERS = [
    "Quote Date",
    "SP",
    "Partner",
    "End User",
    "Country",
    "Brand",
    "Product",
    "Value (AED)",
    "GP (AED)",
    "Contact Name",
    "Closure",
    "Probability",
    "Status",
    "Details",
]


def _clean(value: object | None) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value).strip()


def _parse_quote_date(value: str | None) -> datetime | None:
    if not value or not str(value).strip():
        return None
    raw = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d-%b-%y", "%d-%b-%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(raw, fmt).replace(tzinfo=UTC)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(raw.replace("Z", "+00:00"))
    except ValueError:
        return None


def _format_quote_date(value: datetime | None) -> str | None:
    if not value:
        return None
    if value.tzinfo is None:
        value = value.replace(tzinfo=UTC)
    return value.date().isoformat()


def _format_probability(value: object | None) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, float):
        pct = int(round(value * 100)) if value <= 1 else int(round(value))
        return f"{pct}%"
    text = str(value).strip()
    if not text:
        return ""
    if text.endswith("%"):
        return text
    try:
        num = float(text)
        if num <= 1:
            return f"{int(round(num * 100))}%"
        return f"{int(round(num))}%"
    except ValueError:
        return text


def _format_money(value: object | None) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, (int, float)):
        return f"{value:,.2f}".rstrip("0").rstrip(".") if isinstance(value, float) else f"{value:,}"
    text = str(value).strip().replace("AED", "").replace(",", "").strip()
    try:
        num = float(text)
        if num == int(num):
            return f"{int(num):,}"
        return f"{num:,.2f}"
    except ValueError:
        return str(value).strip()


def to_pipeline_out(row: PipelineEntry, viewer: User | None = None) -> PipelineEntryOut:
    from app.activity_log import logs_for_viewer

    return PipelineEntryOut(
        id=row.id,
        quote_date=_format_quote_date(row.quote_date),
        sp=row.sp or "",
        partner=row.partner or "",
        end_user=row.end_user or "",
        country=row.country or "",
        brand=row.brand or "",
        product=row.product or "",
        value_aed=row.value_aed or "",
        gp_aed=row.gp_aed or "",
        contact_name=row.contact_name or "",
        closure=row.closure or "",
        probability=row.probability or "",
        status=row.status or "",
        details=row.details or "",
        activity_logs=logs_for_viewer(row, viewer, fallback_created_action="Created pipeline entry"),
        created_at=iso(row.created_at) or "",
        updated_at=iso(row.updated_at) or "",
    )


def _matches_multi(value: str | None, filter_value: str | None) -> bool:
    """Match a field against a single value or comma-separated multi-select filter."""
    if not filter_value or filter_value == "all":
        return True
    options = {part.strip().lower() for part in filter_value.split(",") if part.strip()}
    if not options:
        return True
    return (value or "").strip().lower() in options


def _filter_pipeline_rows(
    rows: list[PipelineEntry],
    search: str | None = None,
    status_filter: str | None = None,
    brand: str | None = None,
    country: str | None = None,
    sp: str | None = None,
    closure: str | None = None,
) -> list[PipelineEntry]:
    term = (search or "").strip().lower()
    filtered: list[PipelineEntry] = []
    for row in rows:
        if not _matches_multi(row.status, status_filter):
            continue
        if not _matches_multi(row.brand, brand):
            continue
        if not _matches_multi(row.country, country):
            continue
        if not _matches_multi(row.sp, sp):
            continue
        if not _matches_multi(row.closure, closure):
            continue
        if term:
            haystack = " ".join([
                row.partner or "",
                row.end_user or "",
                row.product or "",
                row.contact_name or "",
                row.brand or "",
                row.status or "",
                row.details or "",
                row.sp or "",
                row.country or "",
            ]).lower()
            if term not in haystack:
                continue
        filtered.append(row)
    return filtered


def query_pipeline(
    db: Session,
    search: str | None = None,
    status_filter: str | None = None,
    brand: str | None = None,
    country: str | None = None,
    sp: str | None = None,
    closure: str | None = None,
    user: User | None = None,
) -> PipelineListResponse:
    rows = db.query(PipelineEntry).order_by(PipelineEntry.quote_date.desc().nulls_last(), PipelineEntry.id.desc()).all()
    filtered = _filter_pipeline_rows(rows, search, status_filter, brand, country, sp, closure)
    return PipelineListResponse(items=[to_pipeline_out(r, user) for r in filtered], total=len(filtered))


def create_pipeline_entry(db: Session, payload: PipelineEntryCreate, user: User) -> PipelineEntryOut:
    from app.activity_log import append_activity

    row = PipelineEntry(
        quote_date=_parse_quote_date(payload.quote_date),
        sp=_clean(payload.sp),
        partner=_clean(payload.partner),
        end_user=_clean(payload.end_user),
        country=_clean(payload.country),
        brand=_clean(payload.brand),
        product=_clean(payload.product),
        value_aed=_clean(payload.value_aed),
        gp_aed=_clean(payload.gp_aed),
        contact_name=_clean(payload.contact_name),
        closure=_clean(payload.closure),
        probability=_format_probability(payload.probability),
        status=_clean(payload.status),
        details=_clean(payload.details),
        created_by_id=user.id,
        activity_logs_json="[]",
    )
    append_activity(row, user, "Created pipeline entry", f"{row.partner or 'Entry'} · {row.status or '—'}")
    db.add(row)
    db.commit()
    db.refresh(row)
    return to_pipeline_out(row, user)


def update_pipeline_entry(
    db: Session,
    entry_id: int,
    payload: PipelineEntryUpdate,
    user: User | None = None,
) -> PipelineEntryOut:
    from app.activity_log import append_activity

    row = db.query(PipelineEntry).filter(PipelineEntry.id == entry_id).first()
    if not row:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Pipeline entry not found")

    data = payload.model_dump(exclude_unset=True)
    if "quote_date" in data:
        row.quote_date = _parse_quote_date(data["quote_date"])
    for field in (
        "sp", "partner", "end_user", "country", "brand", "product",
        "value_aed", "gp_aed", "contact_name", "closure", "status", "details",
    ):
        if field in data and data[field] is not None:
            setattr(row, field, _clean(data[field]))
    if "probability" in data and data["probability"] is not None:
        row.probability = _format_probability(data["probability"])

    detail = f"Status: {row.status or '—'}" if row.status else "Fields updated"
    append_activity(row, user, "Updated pipeline entry", detail)
    db.add(row)
    db.commit()
    db.refresh(row)
    return to_pipeline_out(row, user)


def delete_pipeline_entry(db: Session, entry_id: int) -> None:
    row = db.query(PipelineEntry).filter(PipelineEntry.id == entry_id).first()
    if not row:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Pipeline entry not found")
    db.delete(row)
    db.commit()


def _excel_cell_value(row: PipelineEntry, header: str):
    mapping = {
        "Quote Date": row.quote_date.date() if row.quote_date else None,
        "SP": row.sp or "",
        "Partner": row.partner or "",
        "End User": row.end_user or "",
        "Country": row.country or "",
        "Brand": row.brand or "",
        "Product": row.product or "",
        "Value (AED)": _money_number(row.value_aed),
        "GP (AED)": _money_number(row.gp_aed),
        "Contact Name": row.contact_name or "",
        "Closure": row.closure or "",
        "Probability": _probability_number(row.probability),
        "Status": row.status or "",
        "Details": row.details or "",
    }
    return mapping.get(header, "")


def _money_number(value: str | None):
    if not value:
        return None
    text = re.sub(r"[^\d.]", "", str(value).replace(",", ""))
    if not text:
        return None
    try:
        num = float(text)
        return int(num) if num == int(num) else num
    except ValueError:
        return value


def _probability_number(value: str | None):
    if not value:
        return None
    text = str(value).strip().replace("%", "")
    try:
        num = float(text)
        # Excel template stores 0.25 for 25%
        return num / 100 if num > 1 else num
    except ValueError:
        return value


def export_pipeline_excel(
    db: Session,
    search: str | None = None,
    status_filter: str | None = None,
    brand: str | None = None,
    country: str | None = None,
    sp: str | None = None,
    closure: str | None = None,
) -> bytes:
    rows = db.query(PipelineEntry).order_by(PipelineEntry.quote_date.asc().nulls_last(), PipelineEntry.id.asc()).all()
    rows = _filter_pipeline_rows(rows, search, status_filter, brand, country, sp, closure)

    if PIPELINE_TEMPLATE.is_file():
        wb = load_workbook(PIPELINE_TEMPLATE)
        ws = wb.active
        # Clear existing data rows under header (row 3)
        header_row = 3
        if ws.max_row > header_row:
            ws.delete_rows(header_row + 1, ws.max_row - header_row)
        # Map headers from template row 3
        headers_by_col: dict[int, str] = {}
        for col in range(1, ws.max_column + 1):
            val = ws.cell(header_row, col).value
            if val:
                headers_by_col[col] = str(val).strip()

        data_font = Font(name="Calibri", size=10, bold=False)
        data_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for i, row in enumerate(rows):
            excel_row = header_row + 1 + i
            for col, header in headers_by_col.items():
                cell = ws.cell(excel_row, col)
                cell.value = _excel_cell_value(row, header)
                if header == "Quote Date" and isinstance(cell.value, date):
                    cell.number_format = "D-MMM-YY"
                elif header in {"Value (AED)", "GP (AED)"} and isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'
                elif header == "Probability" and isinstance(cell.value, float):
                    cell.number_format = "0%"
                cell.font = data_font
                cell.alignment = data_alignment
                src = ws.cell(header_row, col)
                if src.has_style and src.border:
                    cell.border = copy(src.border)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Pipeline"
        header_fill = PatternFill("solid", fgColor="548235")
        header_font = Font(bold=True, color="FFFFFF")
        thin = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        for col, header in enumerate(HEADERS, start=1):
            cell = ws.cell(1, col, header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin
        data_font = Font(name="Calibri", size=10, bold=False)
        for i, row in enumerate(rows):
            for col, header in enumerate(HEADERS, start=1):
                cell = ws.cell(i + 2, col, _excel_cell_value(row, header))
                cell.border = thin
                cell.font = data_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col in range(1, len(HEADERS) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 16

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _quote_pipeline_status(status: str) -> str:
    s = (status or "").lower()
    if "lost" in s:
        return "Lost"
    if "closed" in s or "delivered" in s:
        return "Won"
    if "pending" in s or "not approved" in s:
        return "On Hold"
    if "submitted" in s or "placed" in s or "approved" in s:
        return "Quoted"
    return "Quoted"


def sync_pipeline_from_quotations(db: Session, user: User) -> PipelineSyncResponse:
    """Upsert open quotations into the pipeline sheet (system-generated review base)."""
    from app.activity_log import append_activity

    quotes = (
        db.query(Quotation)
        .filter(Quotation.phase.in_(["budgetary", "formal"]))
        .order_by(Quotation.updated_at.desc())
        .all()
    )
    created = updated = skipped = 0

    for q in quotes:
        st = (q.status or "").lower()
        if "lost" in st or "closed" in st:
            skipped += 1
            continue

        marker = f"[Q:{q.quote_number}]"
        existing = (
            db.query(PipelineEntry)
            .filter(PipelineEntry.details.contains(marker))
            .first()
        )
        pipe_status = _quote_pipeline_status(q.status)
        details = f"{marker} {q.status}. {(q.details or '').strip()}".strip()
        quote_date = _parse_quote_date(q.quotation_date) if q.quotation_date else None

        if existing:
            existing.quote_date = quote_date or existing.quote_date
            existing.sp = _clean(q.sales_person) or existing.sp
            existing.partner = _clean(q.partner) or existing.partner
            existing.end_user = _clean(q.end_user) or existing.end_user
            existing.country = _clean(q.country) or existing.country
            existing.brand = _clean(q.brand) or existing.brand
            existing.product = _clean(q.products) or existing.product
            existing.value_aed = _clean(q.deal_value) or existing.value_aed
            existing.gp_aed = _clean(q.gp_value) or existing.gp_aed
            existing.contact_name = _clean(q.contact_person) or existing.contact_name
            existing.closure = _clean(q.closure_date) or existing.closure
            existing.probability = _format_probability(q.probability) or existing.probability
            existing.status = pipe_status
            existing.details = details
            append_activity(existing, user, "Synced from quotation", q.quote_number)
            db.add(existing)
            updated += 1
        else:
            row = PipelineEntry(
                quote_date=quote_date,
                sp=_clean(q.sales_person),
                partner=_clean(q.partner),
                end_user=_clean(q.end_user),
                country=_clean(q.country),
                brand=_clean(q.brand),
                product=_clean(q.products),
                value_aed=_clean(q.deal_value),
                gp_aed=_clean(q.gp_value),
                contact_name=_clean(q.contact_person),
                closure=_clean(q.closure_date),
                probability=_format_probability(q.probability),
                status=pipe_status,
                details=details,
                created_by_id=user.id,
                activity_logs_json="[]",
            )
            append_activity(row, user, "Created from quotation sync", q.quote_number)
            db.add(row)
            created += 1

    db.commit()
    total = db.query(PipelineEntry).count()
    return PipelineSyncResponse(
        created=created,
        updated=updated,
        skipped=skipped,
        total_pipeline=total,
        message=f"Synced {created + updated} open quote(s) into pipeline ({skipped} closed/lost skipped).",
    )


def review_pipeline(db: Session, user: User | None = None) -> PipelineListResponse:
    """Pipeline review set: open deals for manager presentation (exclude Lost)."""
    rows = (
        db.query(PipelineEntry)
        .order_by(PipelineEntry.quote_date.desc().nulls_last(), PipelineEntry.id.desc())
        .all()
    )
    review_rows = [
        r for r in rows
        if (r.status or "").strip().lower() not in {"lost"}
    ]
    return PipelineListResponse(
        items=[to_pipeline_out(r, user) for r in review_rows],
        total=len(review_rows),
    )


def seed_pipeline_from_template(db: Session) -> int:
    """Import rows from the Excel template if the pipeline table is empty."""
    if db.query(PipelineEntry).count() > 0:
        return 0
    if not PIPELINE_TEMPLATE.is_file():
        return 0

    wb = load_workbook(PIPELINE_TEMPLATE, data_only=True)
    ws = wb.active
    # Find header row
    header_row = None
    headers: dict[int, str] = {}
    for r in range(1, min(10, ws.max_row) + 1):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(r, c).value
            if val and str(val).strip().lower() == "quote date":
                header_row = r
                break
        if header_row:
            break
    if not header_row:
        return 0

    for c in range(1, ws.max_column + 1):
        val = ws.cell(header_row, c).value
        if val:
            headers[c] = str(val).strip()

    created = 0
    for r in range(header_row + 1, ws.max_row + 1):
        raw: dict[str, object] = {}
        empty = True
        for c, header in headers.items():
            val = ws.cell(r, c).value
            if val is not None and str(val).strip() != "":
                empty = False
            raw[header] = val
        if empty:
            continue

        quote = raw.get("Quote Date")
        quote_date: datetime | None = None
        if isinstance(quote, datetime):
            quote_date = quote.replace(tzinfo=UTC) if quote.tzinfo is None else quote
        elif isinstance(quote, date):
            quote_date = datetime(quote.year, quote.month, quote.day, tzinfo=UTC)
        else:
            quote_date = _parse_quote_date(_clean(quote))

        row = PipelineEntry(
            quote_date=quote_date,
            sp=_clean(raw.get("SP")),
            partner=_clean(raw.get("Partner")),
            end_user=_clean(raw.get("End User")),
            country=_clean(raw.get("Country")),
            brand=_clean(raw.get("Brand")),
            product=_clean(raw.get("Product")),
            value_aed=_format_money(raw.get("Value (AED)")),
            gp_aed=_format_money(raw.get("GP (AED)")),
            contact_name=_clean(raw.get("Contact Name")),
            closure=_clean(raw.get("Closure")),
            probability=_format_probability(raw.get("Probability")),
            status=_clean(raw.get("Status")),
            details=_clean(raw.get("Details")),
        )
        db.add(row)
        created += 1

    if created:
        db.commit()
    return created
