"""Sales Activities — CRUD + Excel export using Sales Activities template.xlsx."""
from __future__ import annotations

from copy import copy
from datetime import date, datetime
from io import BytesIO
from pathlib import Path

from fastapi import HTTPException, status
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.models import SalesActivity, User
from app.schemas import (
    SalesActivityCreate,
    SalesActivityListResponse,
    SalesActivityOut,
    SalesActivityUpdate,
)
from app.security import iso

BACKEND_ROOT = Path(__file__).resolve().parent.parent
PROJECT_ROOT = BACKEND_ROOT.parent
SALES_ACTIVITIES_TEMPLATE = PROJECT_ROOT / "public" / "forms" / "Sales Activities template.xlsx"

HEADERS = [
    "Sales Person",
    "Customer Name",
    "Meeting Date",
    "Customer Contact  Person",
    "Contact Number",
    "Meeting outputs",
]


def _norm_header(value: object | None) -> str:
    if value is None:
        return ""
    return " ".join(str(value).split()).strip().lower()


HEADER_ALIASES = {
    "sales person": "Sales Person",
    "customer name": "Customer Name",
    "meeting date": "Meeting Date",
    "customer contact person": "Customer Contact  Person",
    "contact number": "Contact Number",
    "meeting outputs": "Meeting outputs",
    "meeting output": "Meeting outputs",
}


def to_sales_activity_out(row: SalesActivity, viewer: User | None = None) -> SalesActivityOut:
    from app.activity_log import logs_for_viewer

    return SalesActivityOut(
        id=row.id,
        sales_person=row.sales_person or "",
        customer_name=row.customer_name or "",
        meeting_date=row.meeting_date or "",
        contact_person=row.contact_person or "",
        contact_number=row.contact_number or "",
        meeting_outputs=row.meeting_outputs or "",
        activity_logs=logs_for_viewer(row, viewer, fallback_created_action="Created sales activity"),
        created_at=iso(row.created_at) or "",
        updated_at=iso(row.updated_at) or "",
        created_by_id=row.created_by_id,
    )


def _matches_multi(value: str | None, filter_value: str | None) -> bool:
    if not filter_value or filter_value == "all":
        return True
    options = {part.strip().lower() for part in filter_value.split(",") if part.strip()}
    if not options:
        return True
    return (value or "").strip().lower() in options


def _filter_rows(
    rows: list[SalesActivity],
    search: str | None,
    sales_person: str | None,
) -> list[SalesActivity]:
    q = (search or "").strip().lower()
    out: list[SalesActivity] = []
    for row in rows:
        if not _matches_multi(row.sales_person, sales_person):
            continue
        if q:
            blob = " ".join([
                row.sales_person, row.customer_name, row.contact_person,
                row.contact_number, row.meeting_outputs, row.meeting_date,
            ]).lower()
            if q not in blob:
                continue
        out.append(row)
    return out


def list_sales_activities(
    db: Session,
    search: str | None = None,
    sales_person: str | None = None,
    user: User | None = None,
) -> SalesActivityListResponse:
    rows = db.query(SalesActivity).order_by(SalesActivity.meeting_date.desc(), SalesActivity.id.desc()).all()
    filtered = _filter_rows(rows, search, sales_person)
    return SalesActivityListResponse(
        items=[to_sales_activity_out(r, user) for r in filtered],
        total=len(filtered),
    )


def create_sales_activity(db: Session, payload: SalesActivityCreate, user: User) -> SalesActivityOut:
    from app.activity_log import append_activity

    row = SalesActivity(
        sales_person=payload.sales_person or user.name,
        customer_name=payload.customer_name,
        meeting_date=payload.meeting_date,
        contact_person=payload.contact_person,
        contact_number=payload.contact_number,
        meeting_outputs=payload.meeting_outputs,
        created_by_id=user.id,
        activity_logs_json="[]",
    )
    append_activity(row, user, "Created sales activity", f"{row.customer_name or 'Meeting'} · {row.meeting_date or '—'}")
    db.add(row)
    db.commit()
    db.refresh(row)
    return to_sales_activity_out(row, user)


def update_sales_activity(
    db: Session,
    activity_id: int,
    payload: SalesActivityUpdate,
    user: User | None = None,
) -> SalesActivityOut:
    from app.activity_log import append_activity

    row = db.query(SalesActivity).filter(SalesActivity.id == activity_id).first()
    if not row:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Sales activity not found")
    data = payload.model_dump(exclude_unset=True)
    for key, value in data.items():
        if value is not None:
            setattr(row, key, value)
    append_activity(row, user, "Updated sales activity", f"{row.customer_name or 'Meeting'}")
    db.add(row)
    db.commit()
    db.refresh(row)
    return to_sales_activity_out(row, user)


def delete_sales_activity(db: Session, activity_id: int) -> None:
    row = db.query(SalesActivity).filter(SalesActivity.id == activity_id).first()
    if not row:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Sales activity not found")
    db.delete(row)
    db.commit()


def _parse_meeting_date(value: str) -> date | str:
    text = (value or "").strip()
    if not text:
        return ""
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d-%b-%Y", "%d-%b-%y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return text


def _excel_cell_value(row: SalesActivity, header: str):
    if header == "Sales Person":
        return row.sales_person or ""
    if header == "Customer Name":
        return row.customer_name or ""
    if header == "Meeting Date":
        return _parse_meeting_date(row.meeting_date)
    if header == "Customer Contact  Person":
        return row.contact_person or ""
    if header == "Contact Number":
        return row.contact_number or ""
    if header == "Meeting outputs":
        return row.meeting_outputs or ""
    return ""


def export_sales_activities_excel(
    db: Session,
    search: str | None = None,
    sales_person: str | None = None,
) -> bytes:
    rows = db.query(SalesActivity).order_by(SalesActivity.meeting_date.asc(), SalesActivity.id.asc()).all()
    rows = _filter_rows(rows, search, sales_person)

    if SALES_ACTIVITIES_TEMPLATE.is_file():
        wb = load_workbook(SALES_ACTIVITIES_TEMPLATE)
        ws = wb.active
        header_row = 1
        # Clear data rows under header
        if ws.max_row and ws.max_row > header_row:
            ws.delete_rows(header_row + 1, ws.max_row - header_row)

        headers_by_col: dict[int, str] = {}
        for col in range(1, (ws.max_column or 1) + 1):
            raw = ws.cell(header_row, col).value
            key = _norm_header(raw)
            if key and key in HEADER_ALIASES:
                headers_by_col[col] = HEADER_ALIASES[key]

        data_font = Font(name="Calibri", size=10, bold=False)
        data_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for i, row in enumerate(rows):
            excel_row = header_row + 1 + i
            for col, header in headers_by_col.items():
                cell = ws.cell(excel_row, col)
                cell.value = _excel_cell_value(row, header)
                if header == "Meeting Date" and isinstance(cell.value, date):
                    cell.number_format = "D-MMM-YY"
                cell.font = data_font
                cell.alignment = data_alignment
                src = ws.cell(header_row, col)
                if src.has_style and src.border:
                    cell.border = copy(src.border)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sales Activities"
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(bold=True, color="FFFFFF")
        thin = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        for col, header in enumerate(HEADERS, start=1):
            cell = ws.cell(1, col, header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin
        for i, row in enumerate(rows):
            for col, header in enumerate(HEADERS, start=1):
                cell = ws.cell(i + 2, col, _excel_cell_value(row, header))
                cell.border = thin
                if header == "Meeting Date" and isinstance(cell.value, date):
                    cell.number_format = "D-MMM-YY"
        for col in range(1, len(HEADERS) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
